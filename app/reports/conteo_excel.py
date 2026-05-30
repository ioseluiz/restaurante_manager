from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile

_RED     = "A20F22"
_GRAY    = "F5F5F5"
_YELLOW  = "FFFDE7"
_BLUE    = "F0F4FF"
_DARK    = "2C3E50"
_BORDER  = "CCCCCC"
_WHITE   = "FFFFFF"
_HINT    = "555555"

_thin = lambda: Side(style="thin", color=_BORDER)

def _border():
    s = _thin()
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font(color=_DARK, bold=False, italic=False, size=10, name="Calibri"):
    return Font(color=color, bold=bold, italic=italic, size=size, name=name)

def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def generar_excel_conteo(conteo_id, fecha, descripcion, categoria_nombre, filas):
    """
    Generate a physical inventory count Excel file and return the file path.

    filas: list of dicts with keys:
        numero          int
        nombre          str
        unidad          str   (base unit name)
        presentaciones  list[str]
    """
    path = os.path.join(tempfile.gettempdir(), f"conteo_inventario_{conteo_id}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Conteo #{conteo_id}"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5    # #
    ws.column_dimensions["B"].width = 38   # Descripción
    ws.column_dimensions["C"].width = 13   # Unidad Base
    ws.column_dimensions["D"].width = 32   # Presentaciones
    ws.column_dimensions["E"].width = 16   # Cant. Física
    ws.column_dimensions["F"].width = 16   # Unidad Usada

    r = 1  # current row pointer

    # ── Row 1: Restaurant name ────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "Restaurante Italos"
    c.font  = _font(color=_RED, bold=True, size=15)
    c.alignment = _align()
    ws.row_dimensions[r].height = 26
    r += 1

    # ── Row 2: Document title ─────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = "Formato de Toma de Inventario Físico"
    c.font  = _font(color=_DARK, bold=True, size=11)
    c.alignment = _align()
    ws.row_dimensions[r].height = 20
    r += 1

    # ── Row 3: Metadata ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    meta = (
        f"Conteo N°: {conteo_id}     "
        f"Fecha: {fecha}     "
        f"Categoría: {categoria_nombre or 'Todas las categorías'}"
    )
    if descripcion:
        meta += f"     Descripción: {descripcion}"
    c = ws[f"A{r}"]
    c.value = meta
    c.font  = _font(color=_DARK, size=9)
    c.alignment = _align()
    ws.row_dimensions[r].height = 16
    r += 1

    # ── Row 4: Instructions ───────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    c = ws[f"A{r}"]
    c.value = (
        "Instrucciones: Cuente el stock real en bodega y complete las columnas "
        "Cant. Física y Unidad usada. Si usa una presentación de compra, anote la "
        "cantidad en esa unidad y escriba el nombre de la presentación en la columna "
        "Unidad usada. Registre decimales si aplica."
    )
    c.font      = _font(color=_HINT, italic=True, size=8)
    c.alignment = _align(h="left")
    ws.row_dimensions[r].height = 34
    r += 1

    # ── Row 5: Table header ───────────────────────────────────────────────────
    headers = [
        "#",
        "Descripción del Insumo",
        "Unidad Base",
        "Presentaciones Disponibles",
        "Cant. Física",
        "Unidad Usada",
    ]
    hdr_row = r
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=r, column=col)
        c.value     = text
        c.font      = _font(color=_WHITE, bold=True, size=9)
        c.fill      = _fill(_RED)
        c.alignment = _align()
        c.border    = _border()
    ws.row_dimensions[r].height = 28
    r += 1

    # Freeze panes below header block
    ws.freeze_panes = ws.cell(row=r, column=1)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, fila in enumerate(filas):
        row_fill = _fill(_GRAY) if i % 2 == 0 else _fill(_WHITE)
        pres_list = fila.get("presentaciones") or []
        pres_text = "\n".join(pres_list) if pres_list else "—"
        row_h = max(18, 14 * len(pres_list)) if pres_list else 18

        values = [
            fila["numero"],
            fila["nombre"],
            fila["unidad"],
            pres_text,
            "",   # Cant. Física — left blank for manual entry
            "",   # Unidad Usada — left blank for manual entry
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col)
            c.value     = val
            c.font      = _font(size=9)
            c.border    = _border()
            c.alignment = _align(h="left" if col in (2, 4) else "center")
            if col == 5:
                c.fill = _fill(_YELLOW)
            elif col == 6:
                c.fill = _fill(_BLUE)
            else:
                c.fill = row_fill

        ws.row_dimensions[r].height = row_h
        r += 1

    # ── Signature section ─────────────────────────────────────────────────────
    r += 1
    for label, start_col in [("Elaborado por:", 1), ("Revisado por:", 4)]:
        end_col = start_col + 2
        ws.merge_cells(
            start_row=r, start_column=start_col,
            end_row=r, end_column=end_col,
        )
        c = ws.cell(row=r, column=start_col)
        c.value = label
        c.font  = _font(size=9)

    r += 1
    for start_col in [1, 4]:
        ws.merge_cells(
            start_row=r, start_column=start_col,
            end_row=r, end_column=start_col + 2,
        )
        c = ws.cell(row=r, column=start_col)
        c.value = "_________________________"
        c.font  = _font(size=9)

    r += 1
    for start_col in [1, 4]:
        ws.merge_cells(
            start_row=r, start_column=start_col,
            end_row=r, end_column=start_col + 2,
        )
        c = ws.cell(row=r, column=start_col)
        c.value = "Nombre y Firma"
        c.font  = _font(color=_HINT, italic=True, size=8)

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 1   # Letter
    ws.page_margins.left      = 0.7
    ws.page_margins.right     = 0.7
    ws.page_margins.top       = 0.75
    ws.page_margins.bottom    = 0.75
    ws.print_title_rows       = f"1:{hdr_row}"

    wb.save(path)
    return path
