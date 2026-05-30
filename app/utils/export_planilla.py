"""Excel and PDF export for the Planilla (payroll) module."""
import os
from datetime import date

# ── Excel ─────────────────────────────────────────────────────────────────────
def export_excel(meta, recargos_cfg, ded_cfg, rows, filepath):
    """
    meta       : dict  {periodo_nombre, fecha_inicio, fecha_fin, sucursal, generado}
    recargos_cfg: list of (tipo_hora, nombre_display, recargo)
    ded_cfg    : list of (concepto, nombre_display, porcentaje, aplica_a)
    rows       : list of dicts with keys:
                 nombre, sucursal, sal_hora,
                 h_reg, h_fest, h_dom, h_exd, h_exn,
                 bruto, ss_c, se_c, ded_otras, ded_vales,
                 total_ded, neto, ss_e, se_e, gasto
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter

    RED_HEX   = "A20F22"
    GREEN_HEX = "2E7D32"
    YELLOW_HEX= "F57F17"
    GRAY_HEX  = "EEEEEE"
    BG_GREEN  = "E8F5E9"
    BG_YELLOW = "FFFDE7"
    WHITE     = "FFFFFF"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def money_fmt():
        return '#,##0.00'

    wb = Workbook()

    # ── Sheet 1: Planilla ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Planilla"
    ws.page_setup.orientation = "landscape"
    ws.freeze_panes = "A7"

    # --- Header block (rows 1-4) ---
    ws.merge_cells("A1:S1")
    ws["A1"] = "REPORTE DE PLANILLA"
    ws["A1"].font = Font(bold=True, size=16, color=RED_HEX)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:S2")
    ws["A2"] = f"Período: {meta.get('periodo_nombre','')}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:S3")
    ws["A3"] = (f"Fechas: {meta.get('fecha_inicio','')} — {meta.get('fecha_fin','')}  |  "
                f"Sucursal: {meta.get('sucursal','Todas')}  |  "
                f"Generado: {meta.get('generado', str(date.today()))}")
    ws["A3"].font = Font(size=10, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="center")

    ws["A4"] = ""  # spacer

    # --- Recargos info (row 5) ---
    ws.merge_cells("A5:S5")
    rec_txt = "Recargos aplicados: " + "  |  ".join(
        f"{n}: ×{r}" for _, n, r in recargos_cfg)
    ws["A5"] = rec_txt
    ws["A5"].font = Font(size=9, color="444444", italic=True)
    ws["A5"].alignment = Alignment(horizontal="left")

    # --- Column headers (row 6) ---
    COL_HEADERS = [
        "Empleado", "Sucursal", "Sal/Hora",
        "H. Regulares", "H. Festivos", "H. Domingos", "H. Extra Diurnas", "H. Extra Nocturnas", "Total Horas",
        "Salario Bruto",
        "SS Colaborador", "SE Colaborador",
        "Ded. Otras", "Ded. Vales", "Total Deducciones",
        "Salario Neto",
        "SS Empleador", "SE Empleador", "Gasto Total Empl.",
    ]
    # columns: A=1..S=19
    MONEY_COLS  = {10, 11, 12, 13, 14, 15, 16, 17, 18, 19}  # 1-based
    HOURS_COLS  = {4, 5, 6, 7, 8, 9}
    NETO_COL    = 16
    GASTO_COLS  = {17, 18, 19}

    for ci, hdr in enumerate(COL_HEADERS, 1):
        cell = ws.cell(row=6, column=ci, value=hdr)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = hdr_fill(RED_HEX)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[6].height = 30

    # --- Data rows ---
    for ri, row in enumerate(rows, 7):
        h_reg  = float(row.get("h_reg",  0) or 0)
        h_fest = float(row.get("h_fest", 0) or 0)
        h_dom  = float(row.get("h_dom",  0) or 0)
        h_exd  = float(row.get("h_exd",  0) or 0)
        h_exn  = float(row.get("h_exn",  0) or 0)
        total_h = h_reg + h_fest + h_dom + h_exd + h_exn

        vals = [
            row["nombre"], row["sucursal"], float(row.get("sal_hora", 0) or 0),
            h_reg, h_fest, h_dom, h_exd, h_exn, total_h,
            row["bruto"],
            row["ss_c"], row["se_c"],
            row["ded_otras"], row["ded_vales"], row["total_ded"],
            row["neto"],
            row["ss_e"], row["se_e"], row["gasto"],
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            if ci in MONEY_COLS:
                cell.number_format = money_fmt()
                cell.alignment = Alignment(horizontal="right")
            elif ci in HOURS_COLS:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

            if ci == NETO_COL:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
                cell.font = Font(bold=True, color=GREEN_HEX)
            elif ci in GASTO_COLS:
                cell.fill = PatternFill("solid", fgColor="FFFDE7")
                cell.font = Font(color=YELLOW_HEX)

        # alternate row shading
        if ri % 2 == 0:
            for ci in range(1, len(vals) + 1):
                c = ws.cell(row=ri, column=ci)
                if ci not in (NETO_COL,) and ci not in GASTO_COLS:
                    if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                        c.fill = PatternFill("solid", fgColor="F9F9F9")

    # --- Totals row ---
    if rows:
        tr = 7 + len(rows)
        # text cols
        tc = ws.cell(row=tr, column=1, value="TOTALES")
        tc.font = Font(bold=True, color=WHITE); tc.fill = hdr_fill(RED_HEX)
        tc.alignment = Alignment(horizontal="center"); tc.border = border
        for ci in (2, 3):
            c = ws.cell(row=tr, column=ci, value="")
            c.fill = hdr_fill(RED_HEX); c.border = border

        # hour totals
        for ci in range(4, 10):
            col_letter = get_column_letter(ci)
            first_data = 7
            last_data  = tr - 1
            c = ws.cell(row=tr, column=ci,
                        value=f"=SUM({col_letter}{first_data}:{col_letter}{last_data})")
            c.font = Font(bold=True, color=WHITE)
            c.fill = hdr_fill(RED_HEX)
            c.number_format = "0.00"
            c.alignment = Alignment(horizontal="right")
            c.border = border

        # money totals
        for ci in range(10, 20):
            col_letter = get_column_letter(ci)
            c = ws.cell(row=tr, column=ci,
                        value=f"=SUM({col_letter}7:{col_letter}{tr-1})")
            c.number_format = money_fmt()
            c.alignment = Alignment(horizontal="right")
            c.border = border
            if ci == NETO_COL:
                c.font = Font(bold=True, color=GREEN_HEX)
                c.fill = PatternFill("solid", fgColor="C8E6C9")
            elif ci in GASTO_COLS:
                c.font = Font(bold=True, color=YELLOW_HEX)
                c.fill = PatternFill("solid", fgColor="FFF9C4")
            else:
                c.font = Font(bold=True, color=WHITE)
                c.fill = hdr_fill(RED_HEX)

    # --- Column widths ---
    widths = [28, 16, 10, 12, 12, 12, 14, 15, 11,
              14, 14, 13, 12, 11, 16, 14, 14, 13, 16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    # ── Sheet 2: Configuración ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Configuración")
    ws2["A1"] = "Recargos por tipo de hora"
    ws2["A1"].font = Font(bold=True, size=12, color=RED_HEX)
    ws2.merge_cells("A1:C1")

    for ci, h in enumerate(["Tipo", "Nombre", "Recargo (×)"], 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = hdr_fill(RED_HEX)
        c.border = border

    for ri, (tipo, nombre, recargo) in enumerate(recargos_cfg, 3):
        ws2.cell(row=ri, column=1, value=tipo).border = border
        ws2.cell(row=ri, column=2, value=nombre).border = border
        ws2.cell(row=ri, column=3, value=recargo).border = border

    ws2["A10"] = "Deducciones (%)"
    ws2["A10"].font = Font(bold=True, size=12, color=RED_HEX)
    ws2.merge_cells("A10:D10")

    for ci, h in enumerate(["Concepto", "Nombre", "Porcentaje", "Aplica a"], 1):
        c = ws2.cell(row=11, column=ci, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = hdr_fill(RED_HEX)
        c.border = border

    for ri, (concepto, nombre, pct, aplica) in enumerate(ded_cfg, 12):
        ws2.cell(row=ri, column=1, value=concepto).border = border
        ws2.cell(row=ri, column=2, value=nombre).border = border
        ws2.cell(row=ri, column=3, value=pct).border = border
        ws2.cell(row=ri, column=4, value=aplica).border = border

    for ci, w in enumerate([18, 30, 12, 14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(filepath)


# ── PDF ───────────────────────────────────────────────────────────────────────
def export_pdf(meta, recargos_cfg, ded_cfg, rows, filepath):
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch, cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    RED    = colors.HexColor("#a20f22")
    GREEN  = colors.HexColor("#2e7d32")
    YELLOW = colors.HexColor("#f57f17")
    GRAY   = colors.HexColor("#eeeeee")
    BG_G   = colors.HexColor("#e8f5e9")
    BG_Y   = colors.HexColor("#fffde7")
    WHITE  = colors.white
    DARK   = colors.HexColor("#2c3e50")
    LGRAY  = colors.HexColor("#f5f5f5")

    PAGE_W, PAGE_H = landscape(letter)
    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(letter),
        leftMargin=0.4*inch, rightMargin=0.4*inch,
        topMargin=0.5*inch,  bottomMargin=0.5*inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=16, textColor=RED,
                                 alignment=TA_CENTER, fontName="Helvetica-Bold",
                                 spaceAfter=4)
    sub_style   = ParagraphStyle("sub", fontSize=11, textColor=DARK,
                                 alignment=TA_CENTER, fontName="Helvetica-Bold",
                                 spaceAfter=2)
    info_style  = ParagraphStyle("info", fontSize=8, textColor=colors.HexColor("#666666"),
                                 alignment=TA_CENTER, spaceAfter=6)
    rec_style   = ParagraphStyle("rec", fontSize=7, textColor=colors.HexColor("#444444"),
                                 alignment=TA_LEFT, spaceAfter=8)

    story = []

    # Header
    story.append(Paragraph("REPORTE DE PLANILLA", title_style))
    story.append(Paragraph(f"Período: {meta.get('periodo_nombre','')}", sub_style))
    story.append(Paragraph(
        f"Fechas: {meta.get('fecha_inicio','')} — {meta.get('fecha_fin','')}  &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Sucursal: {meta.get('sucursal','Todas')}  &nbsp;&nbsp;|&nbsp;&nbsp; "
        f"Generado: {meta.get('generado', str(date.today()))}", info_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=RED, spaceAfter=4))

    rec_txt = "Recargos: " + "   |   ".join(f"{n}: ×{r}" for _, n, r in recargos_cfg)
    story.append(Paragraph(rec_txt, rec_style))

    # ── Split into two table sections to fit landscape letter ────────────────
    # Section A: identity + hours (9 cols)
    # Section B: identity + financials (14 cols)

    def fmt_h(v):
        if v is None or v == 0:
            return "0"
        return f"{float(v):.2f}"

    def fmt_m(v):
        if v is None:
            return "$0.00"
        return f"${float(v):,.2f}"

    # ── Section A: Horas ──────────────────────────────────────────────────────
    hdr_a = ["Empleado", "Sucursal", "Sal/Hora",
             "H.Reg", "H.Fest", "H.Dom", "H.ExD", "H.ExN", "Total H."]
    data_a = [hdr_a]
    tot_hreg = tot_hfest = tot_hdom = tot_hexd = tot_hexn = 0.0

    for row in rows:
        h_reg  = float(row.get("h_reg",  0) or 0)
        h_fest = float(row.get("h_fest", 0) or 0)
        h_dom  = float(row.get("h_dom",  0) or 0)
        h_exd  = float(row.get("h_exd",  0) or 0)
        h_exn  = float(row.get("h_exn",  0) or 0)
        tot_hreg += h_reg; tot_hfest += h_fest; tot_hdom += h_dom
        tot_hexd += h_exd; tot_hexn += h_exn
        data_a.append([
            row["nombre"], row["sucursal"], fmt_m(row.get("sal_hora", 0)),
            fmt_h(h_reg), fmt_h(h_fest), fmt_h(h_dom), fmt_h(h_exd), fmt_h(h_exn),
            fmt_h(h_reg + h_fest + h_dom + h_exd + h_exn),
        ])

    if rows:
        data_a.append([
            "TOTALES", "", "",
            fmt_h(tot_hreg), fmt_h(tot_hfest), fmt_h(tot_hdom),
            fmt_h(tot_hexd), fmt_h(tot_hexn),
            fmt_h(tot_hreg + tot_hfest + tot_hdom + tot_hexd + tot_hexn),
        ])

    col_w_a = [2.1*inch, 1.1*inch, 0.75*inch,
               0.65*inch, 0.65*inch, 0.65*inch, 0.75*inch, 0.75*inch, 0.75*inch]

    ts_a = _base_table_style(RED, WHITE, GRAY, LGRAY, len(rows))
    # right-align numeric cols (3..8 = salary + hours)
    for ci in range(2, 9):
        ts_a.add("ALIGN", (ci, 0), (ci, -1), "RIGHT")

    tbl_a = Table(data_a, colWidths=col_w_a, repeatRows=1)
    tbl_a.setStyle(ts_a)

    story.append(Spacer(1, 4))
    story.append(Paragraph("<b>Detalle de Horas</b>",
                            ParagraphStyle("sh", fontSize=9, textColor=RED,
                                           fontName="Helvetica-Bold", spaceAfter=3)))
    story.append(tbl_a)
    story.append(Spacer(1, 12))

    # ── Section B: Cálculos financieros ──────────────────────────────────────
    hdr_b = ["Empleado", "Sucursal",
             "Bruto", "SS Colab.", "SE Colab.",
             "Ded. Otras", "Ded. Vales", "Total Ded.",
             "Neto", "SS Empl.", "SE Empl.", "Gasto Total"]
    data_b = [hdr_b]
    totals = {k: 0.0 for k in
              ("bruto","ss_c","se_c","ded_otras","ded_vales","total_ded","neto","ss_e","se_e","gasto")}

    for row in rows:
        for k in totals:
            totals[k] += float(row.get(k, 0) or 0)
        data_b.append([
            row["nombre"], row["sucursal"],
            fmt_m(row["bruto"]),
            fmt_m(row["ss_c"]), fmt_m(row["se_c"]),
            fmt_m(row["ded_otras"]), fmt_m(row["ded_vales"]), fmt_m(row["total_ded"]),
            fmt_m(row["neto"]),
            fmt_m(row["ss_e"]), fmt_m(row["se_e"]), fmt_m(row["gasto"]),
        ])

    if rows:
        data_b.append([
            "TOTALES", "",
            fmt_m(totals["bruto"]),
            fmt_m(totals["ss_c"]), fmt_m(totals["se_c"]),
            fmt_m(totals["ded_otras"]), fmt_m(totals["ded_vales"]), fmt_m(totals["total_ded"]),
            fmt_m(totals["neto"]),
            fmt_m(totals["ss_e"]), fmt_m(totals["se_e"]), fmt_m(totals["gasto"]),
        ])

    col_w_b = [2.1*inch, 1.0*inch,
               0.82*inch, 0.82*inch, 0.82*inch,
               0.82*inch, 0.82*inch, 0.82*inch,
               0.9*inch,
               0.82*inch, 0.82*inch, 0.9*inch]

    neto_col  = 8   # 0-based in data_b
    gasto_cols = (9, 10, 11)
    tot_row   = len(data_b) - 1

    ts_b = _base_table_style(RED, WHITE, GRAY, LGRAY, len(rows))
    # right-align money cols
    for ci in range(2, 12):
        ts_b.add("ALIGN", (ci, 0), (ci, -1), "RIGHT")
    # Neto column — green bg
    ts_b.add("BACKGROUND", (neto_col, 1), (neto_col, tot_row - 1), BG_G)
    ts_b.add("TEXTCOLOR",  (neto_col, 1), (neto_col, tot_row - 1), GREEN)
    ts_b.add("FONTNAME",   (neto_col, 1), (neto_col, tot_row - 1), "Helvetica-Bold")
    # Gasto columns — yellow bg
    for gc in gasto_cols:
        ts_b.add("BACKGROUND", (gc, 1), (gc, tot_row - 1), BG_Y)
        ts_b.add("TEXTCOLOR",  (gc, 1), (gc, tot_row - 1), YELLOW)
    # Totals row overrides
    if rows:
        ts_b.add("TEXTCOLOR",  (neto_col, tot_row), (neto_col, tot_row), GREEN)
        for gc in gasto_cols:
            ts_b.add("TEXTCOLOR", (gc, tot_row), (gc, tot_row), YELLOW)

    tbl_b = Table(data_b, colWidths=col_w_b, repeatRows=1)
    tbl_b.setStyle(ts_b)

    story.append(Paragraph("<b>Cálculo de Nómina</b>",
                            ParagraphStyle("sh2", fontSize=9, textColor=RED,
                                           fontName="Helvetica-Bold", spaceAfter=3)))
    story.append(tbl_b)

    # ── Page number footer ────────────────────────────────────────────────────
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#999999"))
        canvas.drawString(0.4*inch, 0.3*inch,
                          f"Planilla — {meta.get('periodo_nombre','')} — Página {doc.page}")
        canvas.drawRightString(PAGE_W - 0.4*inch, 0.3*inch,
                               meta.get('generado', str(date.today())))
        canvas.restoreState()

    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)


def _base_table_style(RED, WHITE, GRAY, LGRAY, n_data_rows):
    from reportlab.platypus import TableStyle
    from reportlab.lib import colors

    tot_row = n_data_rows + 1  # header + data; totals = last row if present

    ts = TableStyle([
        # Header
        ("BACKGROUND",  (0, 0), (-1, 0), RED),
        ("TEXTCOLOR",   (0, 0), (-1, 0), WHITE),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUND", (0, 1), (-1, -1),
         [colors.white if i % 2 == 0 else LGRAY for i in range(n_data_rows + 1)]),
        # Totals row
        ("BACKGROUND",  (0, -1), (-1, -1), colors.HexColor("#eeeeee")),
        ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, -1), (-1, -1), 8),
        # All cells
        ("FONTSIZE",    (0, 1), (-1, -2), 7),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ROWHEIGHT",   (0, 0), (-1, -1), 14),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING",   (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
    ])
    return ts
